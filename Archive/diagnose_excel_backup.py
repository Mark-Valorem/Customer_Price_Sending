import pandas as pd
import openpyxl

# This diagnostic script will show us exactly what's in your Excel file
excel_file = r"C:\Users\MarkAnderson\Valorem\Knowledge Hub - Documents\Pricing\Customer Price Lists\Price Sheet Sending_Python\Python_CustomerPricing.xlsx"

print("=== DETAILED EXCEL FILE ANALYSIS ===\n")

# Method 1: Let's see the first 10 rows without any processing
print("1. RAW VIEW - First 10 rows of the Excel file:")
print("-" * 80)
df_raw = pd.read_excel(excel_file, header=None, nrows=10)
for idx, row in df_raw.iterrows():
    print(f"Row {idx}: {row.tolist()}")

print("\n" + "="*80 + "\n")

# Method 2: Let's try different header positions
print("2. TRYING DIFFERENT HEADER POSITIONS:")
print("-" * 80)

for header_row in range(5):  # Try first 5 rows as potential headers
    try:
        df_test = pd.read_excel(excel_file, header=header_row, nrows=3)
        print(f"\nIf we use row {header_row} as header:")
        print(f"Columns found: {df_test.columns.tolist()}")
        if len(df_test) > 0:
            print(f"First data row: {df_test.iloc[0].tolist()}")
    except Exception as e:
        print(f"Error with header={header_row}: {e}")

print("\n" + "="*80 + "\n")

# Method 3: Let's look at the actual Excel structure using openpyxl
print("3. DIRECT EXCEL INSPECTION (using openpyxl):")
print("-" * 80)

wb = openpyxl.load_workbook(excel_file, read_only=True)
sheet = wb.active

print(f"Sheet name: {sheet.title}")
print(f"Max row: {sheet.max_row}")
print(f"Max column: {sheet.max_column}")
print("\nFirst 10 rows (showing first 5 columns only for readability):")

for row_num in range(1, min(11, sheet.max_row + 1)):
    row_data = []
    for col_num in range(1, min(6, sheet.max_column + 1)):
        cell_value = sheet.cell(row=row_num, column=col_num).value
        if cell_value is None:
            row_data.append("EMPTY")
        else:
            row_data.append(str(cell_value)[:30])  # Truncate long values
    print(f"Row {row_num}: {row_data}")

wb.close()

print("\n" + "="*80 + "\n")

# Method 4: Let's find where the actual data starts
print("4. SEARCHING FOR YOUR COLUMN NAMES:")
print("-" * 80)

# Read without headers to search for our expected column names
df_search = pd.read_excel(excel_file, header=None)

# Look for rows containing our expected column names
for idx, row in df_search.iterrows():
    row_str = ' '.join([str(val) for val in row if pd.notna(val)])
    if 'CustomerName' in row_str or 'RecipientName' in row_str:
        print(f"\nFound column names in row {idx}!")
        print(f"Values in this row: {[val for val in row if pd.notna(val)][:8]}")  # Show first 8 non-empty values
        
        # Let's also check what's in the next row
        if idx + 1 < len(df_search):
            next_row = df_search.iloc[idx + 1]
            print(f"Next row contains: {[val for val in next_row if pd.notna(val)][:5]}")  # Show first 5 values